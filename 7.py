
# 把测试结果写入Excel
# def writr_result(filename, sheetname, row, column, final_result):
#     wb = openpyxl.load_workbook(filename)  # 加载工作簿
#     sheet = wb[sheetname]
#     sheet.cell(row=row, column=column).value = final_result  # 写入最终结果
#     wb.save(filename)  # 保存Excel

import openpyxl
import  requests

# 读取Excel测试用例
def read_data(filename,sheetname):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    max_row = sheet.max_row  # 获取sheet里最大的行数
    case_list = []
    for i in range(2, max_row + 1):  # 因为取头不取尾，所以要在最大行数的基础上 +1
        dict1= dict(  # 转行成字典就是为了将测试用例打包成一条一条的数据
        id = sheet.cell(row=i, column=1).value,  # 取的是测试用例来编号id
        url = sheet.cell(row=i, column=5).value,  # 取的是url数据
        data = sheet.cell(row=i, column=6).value,  # 取的是data数据,从Excel取出来的数据都是str
        expect = sheet.cell(row=i, column=7).value)  # 取的是预期结果 （expected数据）
        case_list.append(dict1)
    return case_list  # 设置返回值给别人用

# 发送请求
def api_func(url, requests_body):
    requests_headers_log = {"X-Lemonban-Media-Type": "lemonban.v2",
                            "Content-Type": "application/json"}
    res_log = requests.post(url=url, json=requests_body, headers=requests_headers_log)
    res_log = res_log.json()
    return res_log

# 把测试结果写入到Excel
def writr_result(filename, sheetname, row, column, final_result):
    wb = openpyxl.load_workbook(filename)  # 加载工作簿
    sheet = wb[sheetname]
    sheet.cell(row=row, column=column).value = final_result  # 加入最终结果
    wb.save(filename)

# 执行接口测试，并断言，会写测试结果到Excel
def exeture_func(filename,sheetname):
    res = read_data(filename, sheetname)
    for testcase in res:  # 取出一条一条测试用例
        # 到家之后，把大袋子去掉，把小袋子取出来，把水果从小袋子里一个一个拿出来
        case_id = testcase.get('id')  # 字典取值获取value
        url = testcase.get('url')  # 取出url
        data = testcase.get('data')  # 取出data
        data = eval(data)  # 运行被字符串包裹的Python表达式，用eval把引号去掉
        expect = testcase.get('expect')  # 取出expect，用get取出的都是str
        expect = eval(expect)  # 把字符串转换成字典
        expect_msg = expect.get('msg')  # 从预期结果的字典里吧msg取出来
        # print(case_id, url, data, expect)
        res_1 = api_func(url=url, requests_body=data)  # 调用发送请求的函数，并传入参数
        # print(res_1)
        real_msg = res_1.get('msg')  # 把实际结果取出
        # print(real_msg, real_msg)
        print('预期结果为：{}'.format(expect_msg))
        print('预期结果为：{}'.format(real_msg))
        if real_msg == expect_msg:
            print('这条测试用例执行通过')
            fina_res = '通过了'
        else:
            print('这条测试用例执行不通过')
            fina_res = '不通过有bug'
        print('*' * 100)
        writr_result(filename, sheetname, case_id+1, 8, fina_res)

exeture_func('test_case_api.xlsx', 'register')
exeture_func('test_case_api.xlsx', 'login')









